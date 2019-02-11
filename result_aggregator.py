__author__ = "Nils Rodday"
__copyright__ = "Copyright 2018"
__credits__ = ["Nils Rodday"]
__email__ = "nils.rodday@unibw.de"
__status__ = "Experimental"

import time
import fcntl
import pandas as pd
import openpyxl
from collections import OrderedDict
import operator
import sys
import logging
import os
from subprocess import call

#
#
#This file is implementing an aggregation function via a Queue.
#All temporary input files are being read and stored in the in-memory object of the file.
#Once the queue is empty, the object is being written to a file and sorted.
#
#


def merge_files(pcap_filename, output_folder, workbook, filename):

    #logging.debug('Entered Merge Function ')

    master_ws = workbook[filename]

    tmp_wb = openpyxl.load_workbook(output_folder + 'tmp/' + pcap_filename + '/' +  filename + '.xlsx')
    tmp_ws = tmp_wb[filename]

    #
    #This part is for the first worksheet
    #

    nested_dict_tmp = {}

    #Read file and load to memory in nested dict object
    for cell in tmp_ws['A']:

        #Avoid header line
        if cell.row == 1:
            continue

        if not cell.value in nested_dict_tmp:
            nested_dict_tmp[cell.value] = {}

        if not tmp_ws.cell(row=cell.row, column=2).value in nested_dict_tmp[cell.value]:
            nested_dict_tmp[cell.value][tmp_ws.cell(row=cell.row, column=2).value] = tmp_ws.cell(row=cell.row, column=3).value

    # Sort dictionary and nested dictionary
    for port, list in nested_dict_tmp.items():
        ordered_dict = OrderedDict(sorted(list.items(), key=operator.itemgetter(0)))
        nested_dict_tmp[port] = ordered_dict

    nested_dict_tmp = OrderedDict(sorted(nested_dict_tmp.items(), key=operator.itemgetter(0)))

    #Iterate over the first column in the worksheet (ports) and insert data if data is present in sorted_nested_dict. Remove items from dict once added.
    column_count = master_ws.max_column

    #Write pcap name to master file
    master_ws.cell(row=1, column=column_count + 1).value = pcap_filename

    #Go through master_ws
    for cell in master_ws['A']:
        lookup_dict = nested_dict_tmp.get(cell.value)
        if lookup_dict != None:
            #pop also removes item from dictionary
            amount = lookup_dict.pop(master_ws.cell(row=cell.row, column=2).value, None)
            if amount != None:
                master_ws.cell(row=cell.row, column=column_count + 1).value = amount

    #Add all remaining items to worksheet
    row_count = master_ws.max_row
    while nested_dict_tmp:
        port, nested_dict = nested_dict_tmp.popitem(last=False)

        while nested_dict:
            key, amount = nested_dict.popitem(last=False)

            master_ws.cell(row=row_count + 1, column=1).value = port
            master_ws.cell(row=row_count + 1, column=2).value = key
            master_ws.cell(row=row_count + 1, column=column_count + 1).value = amount

            row_count += 1


    #
    #This part is for the second worksheet
    #

    master_ws_general = workbook['general']
    tmp_ws_general = tmp_wb['general']

    general_dict = {}

    #Fill tmp dictionary from tmp result file
    for cell in tmp_ws_general['A']:
        #Avoid header line
        if cell.row == 1:
            continue

        if not cell.value in general_dict:
            general_dict[cell.value] = tmp_ws_general.cell(row=cell.row, column=2).value

    #Fill master result file from tmp dictionary
    column_count = master_ws_general.max_column

    #Write pcap name to master file
    master_ws_general.cell(row=1, column=column_count + 1).value = pcap_filename

    for cell in master_ws_general['A']:
        amount = general_dict.pop(cell.value, None)

        if amount != None:
            master_ws_general.cell(row=cell.row, column=column_count + 1).value = amount


    return workbook


def sort_file(output_folder, filename):
    logging.debug('Sorting started: ' + filename)

    #File is still locked from calling function
    xl = pd.ExcelFile(output_folder + filename + ".xlsx")
    df = xl.parse(filename)
    df2 = xl.parse('general')
    if 'dscp' in filename:
        df = df.sort_values(by=['Port', 'DSCP']) #sort 1. for Port, 2nd for DSCP value
        df['DSCP'] = df['DSCP'].apply(lambda x: '{0:0>6}'.format(x))  # keep leading zeros
    if 'ecn' in filename:
        df = df.sort_values(by=['Port', 'ECN'])  # sort 1. for Port, 2nd for DSCP value
        df['ECN'] = df['ECN'].apply(lambda x: '{0:0>2}'.format(x))  # keep leading zeros

    writer = pd.ExcelWriter(output_folder + filename + "_sorted" + ".xlsx")
    df.to_excel(writer, sheet_name=filename, index=False)
    df2.to_excel(writer, sheet_name='general', index=False)
    writer.save()


def result_collector(q, output_folder):
    logging.debug('Result collector started!')

    #Get locks on Master output files
    file1 = open(output_folder + 'port_dscp_nested_dict_IPv4.xlsx')
    fcntl.flock(file1, fcntl.LOCK_EX)
    file2 = open(output_folder + 'port_ecn_nested_dict_IPv4.xlsx')
    fcntl.flock(file2, fcntl.LOCK_EX)
    file3 = open(output_folder + 'port_dscp_nested_dict_IPv6.xlsx')
    fcntl.flock(file3, fcntl.LOCK_EX)
    file4 = open(output_folder + 'port_ecn_nested_dict_IPv6.xlsx')
    fcntl.flock(file4, fcntl.LOCK_EX)

    #Opening master files and store them in memory
    port_dscp_nested_dict_IPv4_workbook = openpyxl.load_workbook(output_folder + 'port_dscp_nested_dict_IPv4.xlsx')
    port_ecn_nested_dict_IPv4_workbook = openpyxl.load_workbook(output_folder + 'port_ecn_nested_dict_IPv4.xlsx')
    port_dscp_nested_dict_IPv6_workbook = openpyxl.load_workbook(output_folder + 'port_dscp_nested_dict_IPv6.xlsx')
    port_ecn_nested_dict_IPv6_workbook = openpyxl.load_workbook(output_folder + 'port_ecn_nested_dict_IPv6.xlsx')



    while True:
        if not q.empty():
            item = q.get()
            logging.debug('Picked item from queue: ' + item)

            if item == 'DONE':
                logging.debug('Queue is DONE!')

                #Save all files to disc
                port_dscp_nested_dict_IPv4_workbook.save(output_folder + 'port_dscp_nested_dict_IPv4.xlsx')
                port_dscp_nested_dict_IPv4_workbook.close()
                port_ecn_nested_dict_IPv4_workbook.save(output_folder + 'port_ecn_nested_dict_IPv4.xlsx')
                port_ecn_nested_dict_IPv4_workbook.close()
                port_dscp_nested_dict_IPv6_workbook.save(output_folder + 'port_dscp_nested_dict_IPv6.xlsx')
                port_dscp_nested_dict_IPv6_workbook.close()
                port_ecn_nested_dict_IPv6_workbook.save(output_folder + 'port_ecn_nested_dict_IPv6.xlsx')
                port_ecn_nested_dict_IPv6_workbook.close()

                #Sort the files in the very end
                sort_file(output_folder, 'port_dscp_nested_dict_IPv4')
                sort_file(output_folder, 'port_ecn_nested_dict_IPv4')
                sort_file(output_folder, 'port_dscp_nested_dict_IPv6')
                sort_file(output_folder, 'port_ecn_nested_dict_IPv6')

                #Release locks
                fcntl.flock(file1, fcntl.LOCK_UN)
                fcntl.flock(file2, fcntl.LOCK_UN)
                fcntl.flock(file3, fcntl.LOCK_UN)
                fcntl.flock(file4, fcntl.LOCK_UN)

                #Delete tmp directory
                if os.path.exists(output_folder + "tmp/"):
                    call(["rm", "-r", output_folder + 'tmp/'])

                return

            pcap_filename = item.split('/')[-1]
            port_dscp_nested_dict_IPv4_workbook = merge_files(pcap_filename, output_folder, port_dscp_nested_dict_IPv4_workbook, 'port_dscp_nested_dict_IPv4')
            port_ecn_nested_dict_IPv4_workbook = merge_files(pcap_filename, output_folder, port_ecn_nested_dict_IPv4_workbook, 'port_ecn_nested_dict_IPv4')
            port_dscp_nested_dict_IPv6_workbook = merge_files(pcap_filename, output_folder, port_dscp_nested_dict_IPv6_workbook, 'port_dscp_nested_dict_IPv6')
            port_ecn_nested_dict_IPv6_workbook = merge_files(pcap_filename, output_folder, port_ecn_nested_dict_IPv6_workbook, 'port_ecn_nested_dict_IPv6')

        else:
            logging.debug('Result collector is going to sleep for 5min')
            time.sleep(300)
            sys.stdout.flush()


