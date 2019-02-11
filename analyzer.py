__author__ = "Nils Rodday"
__copyright__ = "Copyright 2018"
__credits__ = ["Nils Rodday"]
__email__ = "nils.rodday@unibw.de"
__status__ = "Experimental"




#The following output files must be present in the output folder and have the correct layout for this script to run properly. The script "xlsx_file_generator.py" is generating the files in their correct layout.
#
#"sorted_dictionary_source.xlsx"
#"sorted_dictionary_destination.xlsx"
#"sorted_dictionary_dscp_IPv4.xlsx"
#"sorted_dictionary_dscp_IPv6.xlsx"
#"sorted_dictionary_ecn_IPv4.xlsx"
#"sorted_dictionary_ecn_IPv6.xlsx"
#
#
#Filenames are not allowed to contain / chars!
#
#



from scapy.all import *
import numpy
import operator
import sys
import fcntl
import logging
import openpyxl
import pandas as pd
import psutil
import os
from collections import OrderedDict
from openpyxl import Workbook
from subprocess import call
from random import randint
from time import sleep
import pickle


if len(sys.argv) != 3:
    print('Usage:')
    print('analyzer.py PCAP_GZ_FILENAME OUTPUT_FOLDER')
    sys.exit()


#Get ramdom pcap every run
#number = random.randint(1,27)

pcap_filepath = sys.argv[1]
output_folder = sys.argv[2]
#output_folder = 'output/'
#pcap_filename = 'even_smaller_samples.pcap.gz'
#pcap_filename = 'input/' + str(number) + 'even_smaller_samples.pcap'
#pcap_filename = 'test.pcap'
#pcap_filename = 'sample.pcap'
#pcap_filename = 'DHCPv6.cap'


#Set Logging option
#logging.basicConfig(format='%(asctime)s %(levelname)-8s %(message)s', filename=pcap_filepath + '.log',level=logging.INFO)


counter_IPv4_pkts = 0
counter_IPv6_pkts = 0
counter_other_pkts = 0
counter_tcp_pkts= 0
counter_tcp_pkts_over_IPv4= 0
counter_tcp_pkts_over_IPv6= 0
counter_udp_pkts= 0
counter_udp_pkts_over_IPv4= 0
counter_udp_pkts_over_IPv6= 0
counter_icmp_pkts= 0
counter_icmp_pkts_over_IPv4= 0
counter_icmp_pkts_over_IPv6= 0
counter_other_proto_pkts= 0
counter_other_proto_pkts_over_IPv4= 0
counter_other_proto_pkts_over_IPv6= 0
counter_total_pkts= 0
counter_dscp_values_IPv4 = []
counter_dscp_values_IPv6 = []
counter_dscp_zero_values_IPv4 = 0
counter_dscp_zero_values_IPv6 = 0
counter_ecn_values_IPv4 = []
counter_ecn_values_IPv6 = []
counter_ecn_zero_values_IPv4 = 0
counter_ecn_zero_values_IPv6 = 0
counter_tos = 0
counter_tos_zero_values_IPv4 = 0
counter_tc_zero_values_IPv6 = 0
counter_port_occurences_source = []
counter_port_occurences_destination = []
port_dscp_nested_dict_IPv4 = {}
port_ecn_nested_dict_IPv4 = {}
port_dscp_nested_dict_IPv6 = {}
port_ecn_nested_dict_IPv6 = {}



def write_ports_to_xlsx(sorted_dict, filename):

    # lock output file (no other process should write meanwhile)
    file = open(output_folder + filename + ".xlsx")
    fcntl.flock(file, fcntl.LOCK_EX)

    # Create a workbook and add a worksheet.
    wb = openpyxl.load_workbook(output_folder + filename + '.xlsx')
    ws = wb[filename]

    #Write new results as new column to file and add pcap name as header
    column_count = ws.max_column
    ws.cell(row=1, column=column_count+1).value = pcap_filename

    # Write total amount of IP packets to file
    ws.cell(row=2, column=column_count + 1).value = counter_IPv4_pkts + counter_IPv6_pkts
    ws.cell(row=3, column=column_count + 1).value = counter_IPv4_pkts
    ws.cell(row=4, column=column_count + 1).value = counter_IPv6_pkts
    ws.cell(row=5, column=column_count + 1).value = counter_other_pkts

    #Write total amount of TCP and UDP packets to file
    ws.cell(row=6, column=column_count + 1).value = counter_tcp_pkts
    ws.cell(row=7, column=column_count + 1).value = counter_tcp_pkts_over_IPv4
    ws.cell(row=8, column=column_count + 1).value = counter_tcp_pkts_over_IPv6
    ws.cell(row=9, column=column_count + 1).value = counter_udp_pkts
    ws.cell(row=10, column=column_count + 1).value = counter_udp_pkts_over_IPv4
    ws.cell(row=11, column=column_count + 1).value = counter_udp_pkts_over_IPv6
    ws.cell(row=12, column=column_count + 1).value = counter_icmp_pkts
    ws.cell(row=13, column=column_count + 1).value = counter_icmp_pkts_over_IPv4
    ws.cell(row=14, column=column_count + 1).value = counter_icmp_pkts_over_IPv6
    ws.cell(row=15, column=column_count + 1).value = counter_other_proto_pkts
    ws.cell(row=16, column=column_count + 1).value = counter_other_proto_pkts_over_IPv4
    ws.cell(row=17, column=column_count + 1).value = counter_other_proto_pkts_over_IPv6

    #Iterate over the data and insert it into the worksheet if cell is present
    for key, value in (sorted_dict):
        #ws.rows[key].value = value
        ws.cell(row=key+18, column=column_count+1).value = value

    for x in range(19, 65554):
        if ws.cell(row=x, column=column_count+1).value == None:
            ws.cell(row=x, column=column_count+1).value = 0

    wb.save(output_folder + filename + '.xlsx')
    wb.close()

    #unlock file
    fcntl.flock(file, fcntl.LOCK_UN)



def write_dscp_to_xlsx(sorted_dict, filename):

    # lock output file (no other process should write meanwhile)
    file = open(output_folder + filename + ".xlsx")
    fcntl.flock(file, fcntl.LOCK_EX)

    # Create a workbook and add a worksheet.
    wb = openpyxl.load_workbook(output_folder + filename + '.xlsx')
    ws = wb[filename]

    #Write new results as new column to file and add pcap name as header
    column_count = ws.max_column
    ws.cell(row=1, column=column_count+1).value = pcap_filename

    # Write total amount of DSCP values to file
    if 'IPv4' in filename:
        ws.cell(row=2, column=column_count + 1).value = len(counter_dscp_values_IPv4)
        ws.cell(row=3, column=column_count + 1).value = counter_dscp_zero_values_IPv4
    elif 'IPv6' in filename:
        ws.cell(row=2, column=column_count + 1).value = len(counter_dscp_values_IPv6)
        ws.cell(row=3, column=column_count + 1).value = counter_dscp_zero_values_IPv6

    #Iterate over the data and insert it into the worksheet if cell is present
    for key, value in (sorted_dict):
        #ws.rows[key].value = value
        for col_cells in ws.iter_cols(min_col=1, max_col=1):
            for cell in col_cells:
                if key == cell.value:
                    ws.cell(row=cell.row, column=column_count+1).value = value

    #Fill remaining empty fields with 0 values
    for col_cells in ws.iter_cols(min_col=column_count+1, max_col=column_count+1):
         for cell in col_cells:
            if cell.value == None and cell.row >=5:
                ws.cell(row=cell.row, column=column_count+1).value = 0

    wb.save(output_folder + filename + '.xlsx')
    wb.close()

    #unlock file
    fcntl.flock(file, fcntl.LOCK_UN)


def write_ecn_to_xlsx(sorted_dict, filename):

    # lock output file (no other process should write meanwhile)
    file = open(output_folder + filename + ".xlsx")
    fcntl.flock(file, fcntl.LOCK_EX)

    # Create a workbook and add a worksheet.
    wb = openpyxl.load_workbook(output_folder + filename + '.xlsx')
    ws = wb[filename]

    #Write new results as new column to file and add pcap name as header
    column_count = ws.max_column
    ws.cell(row=1, column=column_count+1).value = pcap_filename

    # Write total amount of ECN values to file
    if 'IPv4' in filename:
        ws.cell(row=2, column=column_count + 1).value = len(counter_ecn_values_IPv4)
        ws.cell(row=3, column=column_count + 1).value = counter_ecn_zero_values_IPv4
    elif 'IPv6' in filename:
        ws.cell(row=2, column=column_count + 1).value = len(counter_ecn_values_IPv6)
        ws.cell(row=3, column=column_count + 1).value = counter_ecn_zero_values_IPv6

    #Iterate over the data and insert it into the worksheet if cell is present
    for key, value in (sorted_dict):
        #ws.rows[key].value = value
        for col_cells in ws.iter_cols(min_col=1, max_col=1):
            for cell in col_cells:
                if key == cell.value:
                    ws.cell(row=cell.row, column=column_count+1).value = value

    #Fill remaining empty fields with 0 values
    for col_cells in ws.iter_cols(min_col=column_count+1, max_col=column_count+1):
         for cell in col_cells:
            if cell.value == None and cell.row >=5:
                ws.cell(row=cell.row, column=column_count+1).value = 0

    wb.save(output_folder + filename + '.xlsx')
    wb.close()

    #unlock file
    fcntl.flock(file, fcntl.LOCK_UN)


def write_nested_dict_to_xlsx(sorted_nested_dict, filename):

    #start_time = time.time()
    #print('For file ' + pcap_filename +  'Started creating file and writing to it: ')
    #print('--------------------------')
    #sys.stdout.flush()

    # Create a workbook and add a worksheet.
    wb = Workbook()
    ws = wb.create_sheet(filename)

    if filename == 'port_dscp_nested_dict_IPv4' or filename == 'port_dscp_nested_dict_IPv6':

        ws.cell(row=1, column=1).value = 'Port'
        ws.cell(row=1, column=2).value = 'DSCP'

        ws = wb.create_sheet('general')

        ws.cell(row=1, column=1).value = 'Filename'
        ws.cell(row=2, column=1).value = 'IP Packets'
        ws.cell(row=3, column=1).value = 'IPv4 Packets'
        ws.cell(row=4, column=1).value = 'IPv6 Packets'
        ws.cell(row=5, column=1).value = 'Non-IP Packets'
        ws.cell(row=6, column=1).value = 'TCP Packets'
        ws.cell(row=7, column=1).value = 'TCP Packets over IPv4'
        ws.cell(row=8, column=1).value = 'TCP Packets over IPv6'
        ws.cell(row=9, column=1).value = 'UDP Packets'
        ws.cell(row=10, column=1).value = 'UDP Packets over IPv4'
        ws.cell(row=11, column=1).value = 'UDP Packets over IPv6'
        ws.cell(row=12, column=1).value = 'ICMP Packets'
        ws.cell(row=13, column=1).value = 'ICMP Packets over IPv4'
        ws.cell(row=14, column=1).value = 'ICMP Packets over IPv6'
        ws.cell(row=15, column=1).value = 'Other Packets over IP'
        ws.cell(row=16, column=1).value = 'Other Packets over IPv4'
        ws.cell(row=17, column=1).value = 'Other Packets over IPv6'
        ws.cell(row=18, column=1).value = ''
        ws.cell(row=19, column=1).value = 'Packets with DSCP values'
        ws.cell(row=20, column=1).value = 'Packets without DSCP values'
        ws.cell(row=21, column=1).value = ''
        ws.cell(row=22, column=1).value = 'Packets with entire TOS field to 0 in IPv4'
        ws.cell(row=23, column=1).value = 'Packets with entire TC field to 0 in IPv6'

    elif filename == 'port_ecn_nested_dict_IPv4' or filename == 'port_ecn_nested_dict_IPv6':

        ws.cell(row=1, column=1).value = 'Port'
        ws.cell(row=1, column=2).value = 'ECN'

        ws = wb.create_sheet('general')

        ws.cell(row=1, column=1).value = 'Filename'
        ws.cell(row=2, column=1).value = 'IP Packets'
        ws.cell(row=3, column=1).value = 'IPv4 Packets'
        ws.cell(row=4, column=1).value = 'IPv6 Packets'
        ws.cell(row=5, column=1).value = 'Non-IP Packets'
        ws.cell(row=6, column=1).value = 'TCP Packets'
        ws.cell(row=7, column=1).value = 'TCP Packets over IPv4'
        ws.cell(row=8, column=1).value = 'TCP Packets over IPv6'
        ws.cell(row=9, column=1).value = 'UDP Packets'
        ws.cell(row=10, column=1).value = 'UDP Packets over IPv4'
        ws.cell(row=11, column=1).value = 'UDP Packets over IPv6'
        ws.cell(row=12, column=1).value = 'ICMP Packets'
        ws.cell(row=13, column=1).value = 'ICMP Packets over IPv4'
        ws.cell(row=14, column=1).value = 'ICMP Packets over IPv6'
        ws.cell(row=15, column=1).value = 'Other Packets over IP'
        ws.cell(row=16, column=1).value = 'Other Packets over IPv4'
        ws.cell(row=17, column=1).value = 'Other Packets over IPv6'
        ws.cell(row=18, column=1).value = ''
        ws.cell(row=19, column=1).value = 'Packets with ECN values'
        ws.cell(row=20, column=1).value = 'Packets without ECN values'
        ws.cell(row=21, column=1).value = ''
        ws.cell(row=22, column=1).value = 'Packets with entire TOS field to 0 in IPv4'
        ws.cell(row=23, column=1).value = 'Packets with entire TC field to 0 in IPv6'


    ws = wb['general']

    #Write new results as new column to file and add pcap name as header
    column_count = ws.max_column
    ws.cell(row=1, column=column_count+1).value = pcap_filename

    # Write total amount of IP packets to file
    ws.cell(row=2, column=column_count + 1).value = counter_IPv4_pkts + counter_IPv6_pkts
    ws.cell(row=3, column=column_count + 1).value = counter_IPv4_pkts
    ws.cell(row=4, column=column_count + 1).value = counter_IPv6_pkts
    ws.cell(row=5, column=column_count + 1).value = counter_other_pkts

    #Write total amount of TCP and UDP packets to file
    ws.cell(row=6, column=column_count + 1).value = counter_tcp_pkts
    ws.cell(row=7, column=column_count + 1).value = counter_tcp_pkts_over_IPv4
    ws.cell(row=8, column=column_count + 1).value = counter_tcp_pkts_over_IPv6
    ws.cell(row=9, column=column_count + 1).value = counter_udp_pkts
    ws.cell(row=10, column=column_count + 1).value = counter_udp_pkts_over_IPv4
    ws.cell(row=11, column=column_count + 1).value = counter_udp_pkts_over_IPv6
    ws.cell(row=12, column=column_count + 1).value = counter_icmp_pkts
    ws.cell(row=13, column=column_count + 1).value = counter_icmp_pkts_over_IPv4
    ws.cell(row=14, column=column_count + 1).value = counter_icmp_pkts_over_IPv6
    ws.cell(row=15, column=column_count + 1).value = counter_other_proto_pkts
    ws.cell(row=16, column=column_count + 1).value = counter_other_proto_pkts_over_IPv4
    ws.cell(row=17, column=column_count + 1).value = counter_other_proto_pkts_over_IPv6

    # Write total amount of DSCP/ECN values to file
    if 'dscp' in filename:
        if 'IPv4' in filename:
            ws.cell(row=19, column=column_count + 1).value = len(counter_dscp_values_IPv4)
            ws.cell(row=20, column=column_count + 1).value = counter_dscp_zero_values_IPv4
        elif 'IPv6' in filename:
            ws.cell(row=19, column=column_count + 1).value = len(counter_dscp_values_IPv6)
            ws.cell(row=20, column=column_count + 1).value = counter_dscp_zero_values_IPv6
    if 'ecn' in filename:
        if 'IPv4' in filename:
            ws.cell(row=19, column=column_count + 1).value = len(counter_ecn_values_IPv4)
            ws.cell(row=20, column=column_count + 1).value = counter_ecn_zero_values_IPv4
        elif 'IPv6' in filename:
            ws.cell(row=19, column=column_count + 1).value = len(counter_ecn_values_IPv6)
            ws.cell(row=20, column=column_count + 1).value = counter_ecn_zero_values_IPv6

    #Write Zero TOS and TC values to file
    ws.cell(row=22, column=column_count + 1).value = counter_tos_zero_values_IPv4
    ws.cell(row=23, column=column_count + 1).value = counter_tc_zero_values_IPv6

    #Switch to next worksheet for better aggregation of result values
    ws = wb[filename]
    column_count = ws.max_column

    ws.cell(row=1, column=column_count + 1).value = pcap_filename


    #Add all items to file
    row_count = ws.max_row
    while sorted_nested_dict:
        port, nested_dict = sorted_nested_dict.popitem(last=False)

        while nested_dict:
            key, amount = nested_dict.popitem(last=False)

            ws.cell(row=row_count + 1, column=1).value = port
            ws.cell(row=row_count + 1, column=2).value = key
            ws.cell(row=row_count + 1, column=column_count + 1).value = amount

            row_count += 1


    #middle_time = time.time()
    #temp = middle_time - start_time
    #hours = temp // 3600
    #temp = temp - 3600 * hours
    #minutes = temp // 60
    #seconds = temp - 60 * minutes
    #print('For file ' + pcap_filename +  ' finished processing in memory, now starting to write to disk (Hours:Minutes:Seconds): ' + '%d:%d:%d' % (hours, minutes, seconds))
    #print('--------------------------')
    #sys.stdout.flush()

    #create directory for files
    if not os.path.exists(output_folder + "tmp/" + pcap_filename):
        call(["mkdir", output_folder + 'tmp/' + pcap_filename])

    #Save output to file
    wb.save(output_folder + 'tmp/' + pcap_filename + '/' + filename + ".xlsx")
    wb.close()


    #end_time = time.time()
    #temp = end_time - start_time
    #hours = temp // 3600
    #temp = temp - 3600 * hours
    #minutes = temp // 60
    #seconds = temp - 60 * minutes
    #print('For file ' + pcap_filename +  'reading and writing it took (Hours:Minutes:Seconds): ' + '%d:%d:%d' % (hours, minutes, seconds))
    #print('--------------------------')
    #sys.stdout.flush()




def write_port_and_TOS_to_nested_dict(value, nested_dict, packet):
    l4 = None
    if packet.haslayer(TCP):
        l4 = 'TCP'
    elif packet.haslayer(UDP):
        l4 = 'UDP'

    #Avoid other L4 protocols than UDP and TCP
    if l4 == 'UDP' or l4 == 'TCP':

        #If port is not present, add it

        #Add nested dict in case port is not in dict
        if not packet[l4].sport in nested_dict.keys():
            nested_dict[packet[l4].sport] = []

        #Add DSCP value to port
        nested_dict[packet[l4].sport].append(value)

    return nested_dict


def has_handle(fpath):
    logging.info('Checking if lock is on file')
    logging.info("current file path is : " + fpath)
    foldername = os.path.basename(fpath)
    #logging.info("Directory name is : " + foldername)
    for proc in psutil.process_iter():
        try:
            for item in proc.open_files():
                if fpath == item.path:
                    logging.info('Lock present')
                    return True
        except Exception:
            pass
    logging.info('No Lock present')
    return False


#Go through filenames in output file and check if computation has already been done before
def check_if_file_has_already_been_crunched(pcap_filename, output_folder):

    #lock output file (no other process should write meanwhile)
    file = open(output_folder + 'sorted_dictionary_ecn_IPv6' + '.xlsx')
    fcntl.flock(file, fcntl.LOCK_EX)

    wb = openpyxl.load_workbook(output_folder + 'sorted_dictionary_ecn_IPv6' + '.xlsx', read_only=True)
    ws = wb['general']

    for cell in ws[1]:
        if cell.value == pcap_filename:
            wb.close()
            fcntl.flock(file, fcntl.LOCK_UN)
            return True
    wb.close()
    fcntl.flock(file, fcntl.LOCK_UN)
    return False





#
#
#MAIN PROGRAM
#
#
#

logging.info('Started script')

#Filename is not allowed to contain / chars!
pcap_filename = pcap_filepath.split('/')[-1]
#logging.info('Filename: ' + pcap_filename)


#Make sure that input file has not been processed before
#if check_if_file_has_already_been_crunched(pcap_filename, output_folder):
#    logging.info('The file ' + pcap_filename +' has already been crunched and results are present in output files, e.g.: ' + output_folder + 'port_dscp_nested_dict_IPv4' + '.xlsx')
#    logging.info('Aborting execution!')
#    sys.exit()


#create PCAP reader, only open gzip files in memory
packets = PcapReader(gzip.open(pcap_filepath))

for packet in packets:
    #packet.show()

    # check if IP layer is present to avoid exception
    if packet.haslayer(IP):

        #check for IP version and increase counters
        counter_IPv4_pkts += 1

        if packet[IP].tos: #packet[IP].tos will only be true if it different from 0
            counter_tos += 1

            binary_tos = "{0:08b}".format(packet[IP].tos)
            #logging.info('TOS in Decimal: '  + str(packet[IP].tos))
            #logging.info('DSCP:  ECN:')
            #logging.info(binary_tos[:6] + ' ' + binary_tos[-2:])

            # DSCP counter
            if (binary_tos[:6] != '000000'):
                counter_dscp_values_IPv4.append(binary_tos[:6])
                port_dscp_nested_dict_IPv4 = write_port_and_TOS_to_nested_dict(binary_tos[:6], port_dscp_nested_dict_IPv4, packet)
            else:
                counter_dscp_zero_values_IPv4 += 1

            # ECN counter
            if (binary_tos[-2:] != '00'):
                counter_ecn_values_IPv4.append(binary_tos[-2:])
                port_ecn_nested_dict_IPv4 = write_port_and_TOS_to_nested_dict(binary_tos[-2:], port_ecn_nested_dict_IPv4, packet)
            else:
                counter_ecn_zero_values_IPv4 += 1

        # this happens if TOS field is 0
        else:
            counter_tos_zero_values_IPv4 += 1
            counter_dscp_zero_values_IPv4 += 1
            counter_ecn_zero_values_IPv4 += 1
            #logging.info(packet.show())


    elif packet.haslayer(IPv6):

        counter_IPv6_pkts += 1

        if packet[IPv6].tc:
            counter_tos += 1
            binary_tc = "{0:08b}".format(packet[IPv6].tc)
            #logging.info('TC in Decimal: '  + str(packet[IPv6].tc))
            #logging.info('DSCP:  ECN:')
            #logging.info(binary_tc[:6] + ' ' + binary_tc[-2:])

            # DSCP counter
            if (binary_tc[:6] != '000000'):
                counter_dscp_values_IPv6.append(binary_tc[:6])
                port_dscp_nested_dict_IPv6 = write_port_and_TOS_to_nested_dict(binary_tc[:6], port_dscp_nested_dict_IPv6, packet)
            else:
                counter_dscp_zero_values_IPv6 += 1

            # ECN counter
            if (binary_tc[-2:] != '00'):
                counter_ecn_values_IPv6.append(binary_tc[-2:])
                port_ecn_nested_dict_IPv6 = write_port_and_TOS_to_nested_dict(binary_tc[-2:], port_ecn_nested_dict_IPv6, packet)
            else:
                counter_ecn_zero_values_IPv6 += 1

        # this happens if TC field is 0
        else:
            counter_tc_zero_values_IPv6 += 1
            counter_dscp_zero_values_IPv6 += 1
            counter_ecn_zero_values_IPv6 += 1


        #logging.info(packet.show())


    else:
        #All other packets that do not match to IPv4 or IPv6 get caught here
        counter_other_pkts += 1
        #logging.info(packet.show())


    # Check for protocols: TCP/UDP/ICMP/Others
    if (packet.haslayer(TCP)):
        counter_tcp_pkts += 1

        if packet.haslayer(IP):
            counter_tcp_pkts_over_IPv4 += 1
        elif packet.haslayer(IPv6):
            counter_tcp_pkts_over_IPv6 += 1

        # logging.info(packet[TCP].sport)
        # logging.info(packet[TCP].dport)

        counter_port_occurences_source.append(packet[TCP].sport)
        counter_port_occurences_destination.append(packet[TCP].dport)

    elif (packet.haslayer(UDP)):
        counter_udp_pkts += 1
        # logging.info(packet[TCP].sport)
        # logging.info(packet[UDP].dport)

        if packet.haslayer(IP):
            counter_udp_pkts_over_IPv4 += 1
        elif packet.haslayer(IPv6):
            counter_udp_pkts_over_IPv6 += 1

        counter_port_occurences_source.append(packet[UDP].sport)
        counter_port_occurences_destination.append(packet[UDP].dport)

    elif (packet.haslayer(ICMP)):
        counter_icmp_pkts += 1

        if packet.haslayer(IP):
            counter_icmp_pkts_over_IPv4 += 1
        elif packet.haslayer(IPv6):
            counter_icmp_pkts_over_IPv6 += 1

    else:
        counter_other_proto_pkts += 1

        if packet.haslayer(IP):
            counter_other_proto_pkts_over_IPv4 += 1
        elif packet.haslayer(IPv6):
            counter_other_proto_pkts_over_IPv6 += 1



#Sort src ports, then count their occurence
counter_port_occurences_source = numpy.sort(counter_port_occurences_source)
unique, counts = numpy.unique(counter_port_occurences_source, return_counts=True)
dictionary_source = dict(zip(unique, counts))
sorted_dictionary_source = sorted(dictionary_source.items(), key=operator.itemgetter(0))

#Sort dest ports, then count their occurence
counter_port_occurences_destination = numpy.sort(counter_port_occurences_destination)
unique, counts = numpy.unique(counter_port_occurences_destination, return_counts=True)
dictionary_destination = dict(zip(unique, counts))
sorted_dictionary_destination = sorted(dictionary_destination.items(), key=operator.itemgetter(0))

#Sort DSCP values in IPv4, then count their occurence
counter_dscp_values_IPv4 = numpy.sort(counter_dscp_values_IPv4)
unique, counts = numpy.unique(counter_dscp_values_IPv4, return_counts=True)
dictionary_dscp_IPv4 = dict(zip(unique, counts))
sorted_dictionary_dscp_IPv4 = sorted(dictionary_dscp_IPv4.items(), key=operator.itemgetter(0))

#Sort ECN values in IPv4, then count their occurence
counter_ecn_values_IPv4 = numpy.sort(counter_ecn_values_IPv4)
unique, counts = numpy.unique(counter_ecn_values_IPv4, return_counts=True)
dictionary_ecn_IPv4 = dict(zip(unique, counts))
sorted_dictionary_ecn_IPv4 = sorted(dictionary_ecn_IPv4.items(), key=operator.itemgetter(0))

#Sort DSCP values in IPv6, then count their occurence
counter_dscp_values_IPv6 = numpy.sort(counter_dscp_values_IPv6)
unique, counts = numpy.unique(counter_dscp_values_IPv6, return_counts=True)
dictionary_dscp_IPv6 = dict(zip(unique, counts))
sorted_dictionary_dscp_IPv6 = sorted(dictionary_dscp_IPv6.items(), key=operator.itemgetter(0))

#Sort ECN values in IPv6, then count their occurence
counter_ecn_values_IPv6 = numpy.sort(counter_ecn_values_IPv6)
unique, counts = numpy.unique(counter_ecn_values_IPv6, return_counts=True)
dictionary_ecn_IPv6 = dict(zip(unique, counts))
sorted_dictionary_ecn_IPv6 = sorted(dictionary_ecn_IPv6.items(), key=operator.itemgetter(0))


#Sort DSCP values in IPv4 port-dscp map, count their occurences and create nested dict
for port, list in port_dscp_nested_dict_IPv4.items():
    list = numpy.sort(list)
    unique, counts = numpy.unique(list, return_counts=True)
    tmp = dict(zip(unique, counts))
    ordered_dict = OrderedDict(sorted(tmp.items(), key=operator.itemgetter(0)))
    port_dscp_nested_dict_IPv4[port] = ordered_dict

#Sort DSCP values in IPv4 port-dscp map, count their occurences and create nested dict
for port, list in port_ecn_nested_dict_IPv4.items():
    list = numpy.sort(list)
    unique, counts = numpy.unique(list, return_counts=True)
    tmp = dict(zip(unique, counts))
    ordered_dict = OrderedDict(sorted(tmp.items(), key=operator.itemgetter(0)))
    port_ecn_nested_dict_IPv4[port] = ordered_dict

#Sort DSCP values in IPv4 port-dscp map, count their occurences and create nested dict
for port, list in port_dscp_nested_dict_IPv6.items():
    list = numpy.sort(list)
    unique, counts = numpy.unique(list, return_counts=True)
    tmp = dict(zip(unique, counts))
    ordered_dict = OrderedDict(sorted(tmp.items(), key=operator.itemgetter(0)))
    port_dscp_nested_dict_IPv6[port] = ordered_dict


#Sort DSCP values in IPv4 port-dscp map, count their occurences and create nested dict
for port, list in port_ecn_nested_dict_IPv6.items():
    list = numpy.sort(list)
    unique, counts = numpy.unique(list, return_counts=True)
    tmp = dict(zip(unique, counts))
    ordered_dict = OrderedDict(sorted(tmp.items(), key=operator.itemgetter(0)))
    port_ecn_nested_dict_IPv6[port] = ordered_dict


#Sort port dictionaries
port_dscp_nested_dict_IPv4 = OrderedDict(sorted(port_dscp_nested_dict_IPv4.items(), key=operator.itemgetter(0)))
port_ecn_nested_dict_IPv4 = OrderedDict(sorted(port_ecn_nested_dict_IPv4.items(), key=operator.itemgetter(0)))
port_dscp_nested_dict_IPv6 = OrderedDict(sorted(port_dscp_nested_dict_IPv6.items(), key=operator.itemgetter(0)))
port_ecn_nested_dict_IPv6 = OrderedDict(sorted(port_ecn_nested_dict_IPv6.items(), key=operator.itemgetter(0)))

logging.debug('-------------------------------------')
logging.debug('-------------------------------------')

logging.debug('IP Pakete:      ' + str(counter_IPv4_pkts + counter_IPv6_pkts))
logging.debug('IPv4 Pakete:    ' + str(counter_IPv4_pkts))
logging.debug('IPv6 Pakete:    ' + str(counter_IPv6_pkts))
logging.debug('Andere Pakete:  ' + str(counter_other_pkts))

logging.debug('-------------------------------------')
logging.debug('-------------------------------------')

logging.debug('TCP Pakete:              ' + str(counter_tcp_pkts))
logging.debug('TCP Pakete over IPv4:    ' + str(counter_tcp_pkts_over_IPv4))
logging.debug('TCP Pakete over IPv6:    ' + str(counter_tcp_pkts_over_IPv6))
logging.debug('UDP Pakete:              ' + str(counter_udp_pkts))
logging.debug('UDP Pakete over IPv4:    ' + str(counter_udp_pkts_over_IPv4))
logging.debug('UDP Pakete over IPv6:    ' + str(counter_udp_pkts_over_IPv6))
logging.debug('ICMP Pakete:             ' + str(counter_icmp_pkts))
logging.debug('ICMP Pakete over IPv4:   ' + str(counter_icmp_pkts_over_IPv4))
logging.debug('ICMP Pakete over IPv6:   ' + str(counter_icmp_pkts_over_IPv6))
logging.debug('Andere Protokoll Pakete:              ' + str(counter_other_proto_pkts))
logging.debug('Andere Protokoll Pakete over IPv4:    ' + str(counter_other_proto_pkts_over_IPv4))
logging.debug('Andere Protokoll Pakete over IPv6:    ' + str(counter_other_proto_pkts_over_IPv6))

logging.debug('-------------------------------------')
logging.debug('-------------------------------------')

logging.debug('TCP Sender ports: ')
logging.debug(sorted_dictionary_source)
logging.debug('TCP Receiver ports: ')
logging.debug(sorted_dictionary_destination)

logging.debug('-------------------------------------')
logging.debug('-------------------------------------')

logging.debug('Anzahl Pakete mit DSCP values in IPv4: ' + str(len(counter_dscp_values_IPv4)))
logging.debug('Anzahl Pakete ohne DSCP values in IPv4: ' + str(counter_dscp_zero_values_IPv4))
logging.debug('Anzahl Pakete mit DSCP values in IPv6: ' + str(len(counter_dscp_values_IPv6)))
logging.debug('Anzahl Pakete ohne DSCP values in IPv6: ' + str(counter_dscp_zero_values_IPv6))
logging.debug('DSCP values in IPv4: ')
logging.debug(sorted_dictionary_dscp_IPv4)
logging.debug('DSCP values in IPv6: ')
logging.debug(sorted_dictionary_dscp_IPv6)

logging.debug('-------------------------------------')

logging.debug('Anzahl Pakete mit ECN values in IPv4: ' + str(len(counter_ecn_values_IPv4)))
logging.debug('Anzahl Pakete ohne ECN values in IPv4: ' + str(counter_ecn_zero_values_IPv4))
logging.debug('Anzahl Pakete mit ECN values in IPv6: ' + str(len(counter_ecn_values_IPv6)))
logging.debug('Anzahl Pakete ohne ECN values in IPv6: ' + str(counter_ecn_zero_values_IPv6))
logging.debug('ECN values in IPv4: ')
logging.debug(sorted_dictionary_ecn_IPv4)
logging.debug('ECN values in IPv6: ')
logging.debug(sorted_dictionary_ecn_IPv6)

logging.debug('-------------------------------------')

logging.debug('Anzahl Pakete mit gesamtem TOS Feld auf 0 in IPv4: ' + str(counter_tos_zero_values_IPv4))
logging.debug('Anzahl Pakete mit gesamtem TC Feld auf 0 in IPv6: ' + str(counter_tc_zero_values_IPv6))

logging.debug('-------------------------------------')
logging.debug('-------------------------------------')

logging.debug('DSCP values in IPv4 pro port: ')
logging.debug(port_dscp_nested_dict_IPv4)
logging.debug('ECN values in IPv4 pro port: ')
logging.debug(port_ecn_nested_dict_IPv4)
logging.debug('DSCP values in IPv6 pro port: ')
logging.debug(port_dscp_nested_dict_IPv6)
logging.debug('ECN values in IPv6 pro port: ')
logging.debug(port_ecn_nested_dict_IPv6)


logging.debug('-------------------------------------')
logging.debug('-------------------------------------')


#Serialization for later use:
#pickle.dump(port_dscp_nested_dict_IPv4, open( "port_dscp_nested_dict_IPv4.p", "wb" ) )
#pickle.dump(port_ecn_nested_dict_IPv4, open( "port_ecn_nested_dict_IPv4.p", "wb" ) )
#pickle.dump(port_dscp_nested_dict_IPv6, open( "port_dscp_nested_dict_IPv6.p", "wb" ) )
#pickle.dump(port_ecn_nested_dict_IPv6, open( "port_ecn_nested_dict_IPv6.p", "wb" ) )

#Writing data to XLSX files
write_ports_to_xlsx(sorted_dictionary_source, "sorted_dictionary_source")
write_ports_to_xlsx(sorted_dictionary_destination, "sorted_dictionary_destination")
write_dscp_to_xlsx(sorted_dictionary_dscp_IPv4, "sorted_dictionary_dscp_IPv4")
write_dscp_to_xlsx(sorted_dictionary_dscp_IPv6, "sorted_dictionary_dscp_IPv6")
write_ecn_to_xlsx(sorted_dictionary_ecn_IPv4, "sorted_dictionary_ecn_IPv4")
write_ecn_to_xlsx(sorted_dictionary_ecn_IPv6, "sorted_dictionary_ecn_IPv6")
write_nested_dict_to_xlsx(port_dscp_nested_dict_IPv4, "port_dscp_nested_dict_IPv4")
write_nested_dict_to_xlsx(port_dscp_nested_dict_IPv6, "port_dscp_nested_dict_IPv6")
write_nested_dict_to_xlsx(port_ecn_nested_dict_IPv4, "port_ecn_nested_dict_IPv4")
write_nested_dict_to_xlsx(port_ecn_nested_dict_IPv6, "port_ecn_nested_dict_IPv6")




logging.info('Finished script')



















