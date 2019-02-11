__author__ = "Nils Rodday"
__copyright__ = "Copyright 2018"
__credits__ = ["Nils Rodday"]
__email__ = "nils.rodday@unibw.de"
__status__ = "Stable"

from openpyxl import Workbook
import sys
from subprocess import call


if len(sys.argv) == 1:
    print('Please specify the output folder!')
    sys.exit()

#get the filename and type to be generated from parameter
#filename = sys.argv[1]
output_folder = sys.argv[1]

options = ["All", "sorted_dictionary_source", "sorted_dictionary_destination", "sorted_dictionary_dscp_IPv4", "sorted_dictionary_dscp_IPv6", "sorted_dictionary_ecn_IPv4", "sorted_dictionary_ecn_IPv6", "port_dscp_nested_dict_IPv4", "port_dscp_nested_dict_IPv6", "port_ecn_nested_dict_IPv4", "port_ecn_nested_dict_IPv6"]

#Check if script was called with argument (by itself), otherwise let user enter number
if len(sys.argv) == 2:
    # Print out your options
    for i in range(len(options)):
        print(str(i+1) + ":", options[i])
    inp = int(input("Please specify the option: "))
else:
    inp = int(sys.argv[2])


if inp in range(1, 12):
    filename = options[inp-1]
    if filename == "All":
        for i in (range(2, 12)):
            call(["python3", "xlsx_file_generator.py", output_folder, str(i)])
        sys.exit()
else:
    print("Invalid input!")
    sys.exit()


print('-------------------------------------')
print('Working on it...Please wait!')
print('Creating file: ' + filename)
print('-------------------------------------')


# Create a workbook and add a worksheet.
wb = Workbook()
ws = wb.create_sheet(filename)
#wb.remove_sheet('Sheet')

if filename=="All" or filename=='sorted_dictionary_source' or filename=='sorted_dictionary_destination':
    ws.cell(row=1, column=1).value = 'Filename'
    ws.cell(row=2, column=1).value = 'IP Packets'
    ws.cell(row=3, column=1).value = 'IPv4 Packets'
    ws.cell(row=4, column=1).value = 'IPv6 Packets'
    ws.cell(row=5, column=1).value = 'Other Packets'
    ws.cell(row=6, column=1).value = 'TCP Packets'
    ws.cell(row=7, column=1).value = 'TCP Packets over IPv4'
    ws.cell(row=8, column=1).value = 'TCP Packets over IPv6'
    ws.cell(row=9, column=1).value = 'UDP Packets'
    ws.cell(row=10, column=1).value = 'UDP Packets over IPv4'
    ws.cell(row=11, column=1).value = 'UDP Packets over IPv6'
    ws.cell(row=12, column=1).value = 'ICMP Packets'
    ws.cell(row=13, column=1).value = 'ICMP Packets over IPv4'
    ws.cell(row=14, column=1).value = 'ICMP Packets over IPv6'
    ws.cell(row=15, column=1).value = 'Other Packets'
    ws.cell(row=16, column=1).value = 'Other Packets over IPv4'
    ws.cell(row=17, column=1).value = 'Other Packets over IPv6'
    ws.cell(row=18, column=1).value = ''

    for x in range(1, 65536):
        ws['A' + str(x + 18)] = x


elif filename=="All" or filename=='sorted_dictionary_dscp_IPv4' or filename=='sorted_dictionary_dscp_IPv6':
    ws.cell(row=1, column=1).value = 'Filename'
    ws.cell(row=2, column=1).value = 'Packets with DSCP values'
    ws.cell(row=3, column=1).value = 'Packets without DSCP values'

    for x in range(1, 64):
        ws['A' + str(x + 4)] = "{0:06b}".format(x)


elif filename=="All" or filename=='sorted_dictionary_ecn_IPv4' or filename=='sorted_dictionary_ecn_IPv6':
    ws.cell(row=1, column=1).value = 'Filename'
    ws.cell(row=2, column=1).value = 'Packets with ECN values'
    ws.cell(row=3, column=1).value = 'Packets without ECN values'

    for x in range(1, 4):
        ws['A' + str(x + 4)] = "{0:02b}".format(x)

elif filename=="All" or filename=='port_dscp_nested_dict_IPv4' or filename=='port_dscp_nested_dict_IPv6':

    ws.cell(row=1, column=1).value = 'Port'
    ws.cell(row=1, column=2).value = 'DSCP'


    ws = wb.create_sheet('general')

    ws.cell(row=1, column=1).value = 'Filename'
    ws.cell(row=2, column=1).value = 'IP Packets'
    ws.cell(row=3, column=1).value = 'IPv4 Packets'
    ws.cell(row=4, column=1).value = 'IPv6 Packets'
    ws.cell(row=5, column=1).value = 'Non-IP Packets'
    ws.cell(row=6, column=1).value = 'TCP Packets'
    ws.cell(row=7, column=1).value = 'TCP Packets over IPv4'
    ws.cell(row=8, column=1).value = 'TCP Packets over IPv6'
    ws.cell(row=9, column=1).value = 'UDP Packets'
    ws.cell(row=10, column=1).value = 'UDP Packets over IPv4'
    ws.cell(row=11, column=1).value = 'UDP Packets over IPv6'
    ws.cell(row=12, column=1).value = 'ICMP Packets'
    ws.cell(row=13, column=1).value = 'ICMP Packets over IPv4'
    ws.cell(row=14, column=1).value = 'ICMP Packets over IPv6'
    ws.cell(row=15, column=1).value = 'Other Packets over IP'
    ws.cell(row=16, column=1).value = 'Other Packets over IPv4'
    ws.cell(row=17, column=1).value = 'Other Packets over IPv6'
    ws.cell(row=18, column=1).value = ''
    ws.cell(row=19, column=1).value = 'Packets with DSCP values'
    ws.cell(row=20, column=1).value = 'Packets without DSCP values'
    ws.cell(row=21, column=1).value = ''
    ws.cell(row=22, column=1).value = 'Packets with entire TOS field to 0 in IPv4'
    ws.cell(row=23, column=1).value = 'Packets with entire TC field to 0 in IPv6'

elif filename=="All" or filename=='port_ecn_nested_dict_IPv4' or filename=='port_ecn_nested_dict_IPv6':

    ws.cell(row=1, column=1).value = 'Port'
    ws.cell(row=1, column=2).value = 'ECN'


    ws = wb.create_sheet('general')

    ws.cell(row=1, column=1).value = 'Filename'
    ws.cell(row=2, column=1).value = 'IP Packets'
    ws.cell(row=3, column=1).value = 'IPv4 Packets'
    ws.cell(row=4, column=1).value = 'IPv6 Packets'
    ws.cell(row=5, column=1).value = 'Non-IP Packets'
    ws.cell(row=6, column=1).value = 'TCP Packets'
    ws.cell(row=7, column=1).value = 'TCP Packets over IPv4'
    ws.cell(row=8, column=1).value = 'TCP Packets over IPv6'
    ws.cell(row=9, column=1).value = 'UDP Packets'
    ws.cell(row=10, column=1).value = 'UDP Packets over IPv4'
    ws.cell(row=11, column=1).value = 'UDP Packets over IPv6'
    ws.cell(row=12, column=1).value = 'ICMP Packets'
    ws.cell(row=13, column=1).value = 'ICMP Packets over IPv4'
    ws.cell(row=14, column=1).value = 'ICMP Packets over IPv6'
    ws.cell(row=15, column=1).value = 'Other Packets over IP'
    ws.cell(row=16, column=1).value = 'Other Packets over IPv4'
    ws.cell(row=17, column=1).value = 'Other Packets over IPv6'
    ws.cell(row=18, column=1).value = ''
    ws.cell(row=19, column=1).value = 'Packets with ECN values'
    ws.cell(row=20, column=1).value = 'Packets without ECN values'
    ws.cell(row=21, column=1).value = ''
    ws.cell(row=22, column=1).value = 'Packets with entire TOS field to 0 in IPv4'
    ws.cell(row=23, column=1).value = 'Packets with entire TC field to 0 in IPv6'


#     row_count = 14
#     for x in range(1, 65536):
#         for y in range(1, 64):
#             ws['A' + str(row_count)] = x
#             ws['B' + str(row_count)] = "{0:06b}".format(y)
#             row_count += 1



wb.save(output_folder + filename + '.xlsx')

#Save additional sorted version
if filename=='port_ecn_nested_dict_IPv4' or filename=='port_ecn_nested_dict_IPv6'or filename=='port_dscp_nested_dict_IPv4' or filename=='port_dscp_nested_dict_IPv6':
    wb.save(output_folder + filename + '_sorted' + '.xlsx')


wb.close()